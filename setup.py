import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

def generate_excel_report(df_summary, sortie_path, profil_type="Base vs France"):

    # Création du classeur Excel et ajout des feuilles
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Profil Comparatif"

    # Titre du document
    ws_summary["A1"] = f"Profil - {profil_type} (1/2)"
    ws_summary["A1"].font = Font(size=14, bold=True)
    
    # Remplir la feuille avec les données de df_summary
    for row in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(row)

    # Écrire les données du DataFrame dans la feuille de calcul
    for r_idx, row in enumerate(df_summary.itertuples(index=False), 2):  # Commence à la ligne 2
        for c_idx, value in enumerate(row, 1):  # Commence à la colonne 1
            ws_summary.cell(row=r_idx, column=c_idx, value=value)

    # Création de graphiques pour chaque indicateur
    chart = BarChart()
    chart.type = "col"
    chart.title = "Comparaison des valeurs"

    # Références des données pour toutes les colonnes
    # Supposons que la première colonne est pour les catégories (ex : les noms)
    data = Reference(ws_summary, min_col=2, max_col=len(df_summary.columns), min_row=1, max_row=len(df_summary) + 1)  # Inclure toutes les colonnes
    cats = Reference(ws_summary, min_col=1, min_row=2, max_row=len(df_summary) + 1)  # Les catégories commencent à la ligne 2

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Ajouter le graphique à la feuille de résumé
    ws_summary.add_chart(chart, "E5")
    
    # Sauvegarde du fichier Excel
    wb.save(sortie_path)


def EG_profil(sortie, table, var_age, var_sexe, var_pop1=None, modalite_pop1=None, var_pop2=None, modalite_pop2=None, code_geo=None, zone_geo=None):
    
    # Assurez-vous que le dossier de sortie existe
    os.makedirs(sortie, exist_ok=True)

    # Analyse et génération des profils
    result_df = pd.DataFrame()  # Initialisation du DataFrame de résultat

    if var_pop1 and modalite_pop1:  # Comparaison cible vs totale
        population_cible = table[table[var_pop1] == modalite_pop1]
        population_totale = table
        
        # Calcul des statistiques pour la population cible
        effectif_cible = len(population_cible)
        pourcentage_cible = 100 * effectif_cible / len(table) if len(table) > 0 else 0
        
        # Calcul des statistiques pour la population totale
        effectif_totale = len(population_totale)
        pourcentage_totale = 100 * effectif_totale / len(table) if len(table) > 0 else 0

        # Création d'un DataFrame pour les statistiques
        result_cible_vs_totale = pd.DataFrame({
            'Indicateur': ["Population Cible", "Population Totale"],
            'Effectif': [effectif_cible, effectif_totale],
            'Pourcentage': [pourcentage_cible, pourcentage_totale]
        })

        # Ajout des statistiques par âge pour la population cible
        age_stats_cible = population_cible['age_declare'].describe()  # Statistiques descriptives
        age_stats_totale = population_totale['age_declare'].describe()  # Statistiques descriptives pour la population totale

        # Ajout des résultats d'âge dans le DataFrame
        age_stats_cible_df = pd.DataFrame(age_stats_cible).transpose()
        age_stats_totale_df = pd.DataFrame(age_stats_totale).transpose()
        
        age_stats_cible_df['Indicateur'] = 'Statistiques Age Cible'
        age_stats_totale_df['Indicateur'] = 'Statistiques Age Totale'
        
        # Renommer les colonnes pour éviter les conflits
        age_stats_cible_df.columns = ['Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max', 'Indicateur']
        age_stats_totale_df.columns = ['Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max', 'Indicateur']
        
        # Combiner les résultats
        result_df = pd.concat([result_cible_vs_totale, age_stats_cible_df, age_stats_totale_df], ignore_index=True)
        


        # Enregistrement CSV et génération Excel
        csv_path = os.path.join(sortie, 'profil_cible_vs_totale.csv')
        result_df.to_csv(csv_path, index=False)
        excel_path = os.path.join(sortie, '2016_modele_profil_cible.xlsx')
        generate_excel_report(result_df, excel_path)

    if var_pop1 and modalite_pop1 and var_pop2 and modalite_pop2:  # Comparaison de 2 populations
        population_cible1 = table[table[var_pop1] == modalite_pop1]
        population_cible2 = table[table[var_pop2] == modalite_pop2]
        
        result_cible1_vs_cible2 = {
            'Indicateur': ["Cible 1", "Cible 2"],
            'Effectif': [len(population_cible1), len(population_cible2)],
            'Pourcentage': [100 * len(population_cible1) / len(table), 100 * len(population_cible2) / len(table)]
        }
        
        result_df = pd.DataFrame(result_cible1_vs_cible2)

        # Enregistrement CSV et génération Excel
        csv_path = os.path.join(sortie, 'profil_cible1_vs_cible2.csv')
        result_df.to_csv(csv_path, index=False)
        excel_path = os.path.join(sortie, '2016_modele_profil_cible_vs_cible.xlsx')
        generate_excel_report(result_df, excel_path)

    if code_geo and zone_geo:  # Comparaison géographique
        population_geo = table[table[code_geo] == zone_geo]
        
        result_geo = {
            'Indicateur': ["Population géographique", "Population totale"],
            'Effectif': [len(population_geo), len(table)],
            'Pourcentage': [100 * len(population_geo) / len(table), 100]
        }
        
        result_df = pd.DataFrame(result_geo)

        # Enregistrement CSV et génération Excel
        csv_path = os.path.join(sortie, 'profil_zone_geo.csv')
        result_df.to_csv(csv_path, index=False)
        excel_path = os.path.join(sortie, '2016_modele_profil_zone_geo.xlsx')
        generate_excel_report(result_df, excel_path)

    return result_df
